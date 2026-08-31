#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将采集到的个税 JSON 数据灌入"个税退税计算模板.xlsx"，生成一份填好的新 Excel。

核心原则：
  1. 保留模板里 *所有* 公式（F/K/P/Q 列及 29~42 行的计税公式一律不动）。
  2. 只向"输入区"单元格（淡蓝格）写数：
       - 收入项目：直接写金额到 B / G / L 列；
       - 直接扣除项（养老/医疗/失业/公积金/年金/税延/个人养老金/捐赠/其他）：
         写金额到 B / G / L 列（因其 F=B）；
       - 公式扣除项（子女教育/婴幼儿照护/赡养老人/房贷利息/住房租金）：
         在保持模板固定基数的前提下，求解"月份数"使公式恰好复现采集金额；
       - 已预缴税额（第 40 行）：写"已缴税额"。
  3. 年份表头按实际申报年度升序重排（模板默认 23/24/25 -> 实际 2022/2023/2024）。

用法：
  python3 fill_template.py [数据json] [模板xlsx] [输出xlsx]
  默认：
    数据 = ./tax_data.json
    模板 = ./个税退税计算模板.xlsx
    输出 = ./output/个税退税计算表_周杨_20260802.xlsx
"""

import json
import os
import re
import sys
from copy import copy

import openpyxl

# ----------------------------------------------------------- 路径默认值
HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_JSON = os.path.join(HERE, "tax_data.json")
DEFAULT_TPL = os.path.join(HERE, "个税退税计算模板.xlsx")
DEFAULT_OUT = os.path.join(HERE, "output", "个税退税计算表_已填充.xlsx")

# 三个年度列的"输入列"字母（F/K/P 是公式输出列，不写）
YEAR_COLS = ["B", "G", "L"]          # 收入 & 直接扣除项输入列
MONTH_COLS = ["C", "H", "M"]         # 公式扣除项的"月份"输入列
TYPE_COLS = ["D", "I", "N"]          # 计算类型 / 是否 列
NUM_COLS = ["E", "J", "O"]           # 人数 列
OUT_COLS = ["F", "K", "P"]           # 公式输出列（只读，不写）
PREPAID_COLS = ["F", "K", "P"]       # 第40行 已预缴 是输入格（这3列在40行恰好是输入）
REFUND_COLS = ["F", "K", "P"]        # 第41行 已退税 是输入格


def to_float(v, default=0.0):
    """把 '298254.85' / '0.00' / None 统一转成 float。"""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("元", "")
    if not s or s in ("--", "—", "-"):
        return default
    try:
        return float(s)
    except ValueError:
        return default


def extract_year(申报项目):
    """'2024年度综合所得年度汇算' -> 2024"""
    m = re.search(r"(20\d{2})", 申报项目 or "")
    return int(m.group(1)) if m else None


def load_year_data(data):
    """
    把申报详情按 {year: {...}} 整理出来。
    取每个年度的：逐项金额(_flat)、汇总(已缴税额)、基础信息。
    """
    details = data.get("申报", {}).get("详情", []) or []
    by_year = {}
    for d in details:
        year = extract_year(d.get("_申报项目"))
        if not year:
            continue
        sec = d.get("_detail_sections", {})
        calc = sec.get("计税详情", {}) or {}
        flat = (calc.get("逐项金额", {}) or {}).get("_flat", {}) or {}
        summary = calc.get("汇总", {}) or {}
        base = (sec.get("基础信息", {}) or {}).get("kv", {}) or {}
        by_year[year] = {
            "flat": flat,
            "summary": summary,
            "base": base,
        }
    return by_year


def solve_months(amount, base, cap=12):
    """
    在固定 base（每月标准）和 100% 承担下，求需要几个月才能恰好等于 amount。
    返回整数月份（向上/向下取最近，且不超过 cap）。amount<=0 返回 0。
    用于：子女教育/婴幼儿照护/赡养老人/房贷利息/住房租金 等公式行。
    """
    amount = round(to_float(amount), 2)
    base = to_float(base, 0)
    if amount <= 0 or base <= 0:
        return 0
    months = amount / base
    mi = int(round(months))
    if mi < 0:
        mi = 0
    if mi > cap:
        mi = cap
    return mi


def main(json_path=DEFAULT_JSON, tpl_path=DEFAULT_TPL, out_path=DEFAULT_OUT):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    by_year = load_year_data(data)
    if not by_year:
        sys.exit("✗ JSON 里没解析到任何年度的申报详情")

    years = sorted(by_year.keys())           # 升序
    if len(years) > 3:
        years = years[-3:]                    # 模板只有 3 列，取最近 3 年
    print(f"→ 命中年度：{years}")

    # 取纳税人姓名（取任一年度的基础信息即可，全填同一个人）
    name = ""
    for y in years:
        name = by_year[y]["base"].get("个人基础信息") or ""
        if name:
            break

    wb = openpyxl.load_workbook(tpl_path, data_only=False)   # 保留公式
    ws = wb["个人所得税"]

    # ---- 1. 年份表头重排：B2 / G2 / L2
    for i, y in enumerate(years):
        ws[f"{YEAR_COLS[i]}2"] = f"{y}年"

    # ---- 2. 收入项目（第 4~7 行）写金额到输入列
    income_rows = {
        4: "工资薪金",
        5: "劳务报酬",
        6: "稿酬",
        7: "特许权使用费",
    }
    for i, y in enumerate(years):
        flat = by_year[y]["flat"]
        for r, key in income_rows.items():
            ws[f"{YEAR_COLS[i]}{r}"] = to_float(flat.get(key))

    # ---- 3. 直接扣除项：金额直接写到输入列（F=B 等公式自动取）
    direct_rows = {
        11: "基本养老保险",
        12: "基本医疗保险",
        13: "失业保险",
        14: "住房公积金",
        22: "年金",
        23: "商业健康险",      # F23=B23*12（每月）——写月额度
        24: "税延养老保险",
        25: "个人养老金",
        26: "准予扣除的捐赠额",
        27: "其他",
    }
    for i, y in enumerate(years):
        flat = by_year[y]["flat"]
        for r, key in direct_rows.items():
            val = to_float(flat.get(key))
            if key == "商业健康险":
                val = val / 12.0          # 该行公式是 B*12，所以输入月额度
            ws[f"{YEAR_COLS[i]}{r}"] = val

    # ---- 4. 公式扣除项：保持模板固定基数，求解月份数
    #   行 -> (JSON键, 模板基数单元格, 计算类型值, 是否用人数列)
    formula_rows = {
        15: ("子女教育", 2000, "100%承担", True),    # F=E*B*C
        16: ("3岁以下婴幼儿照护", 2000, "100%承担", True),
        17: ("赡养老人", 3000, "100%承担", False),    # F=B*C（无人数）
        18: ("住房贷款利息", 1000, "是", False),       # F=IF(D="是",B*C,0)
        19: ("住房租金", 1500, "是", False),
    }
    for i, y in enumerate(years):
        flat = by_year[y]["flat"]
        for r, (key, base, type_val, use_num) in formula_rows.items():
            amount = to_float(flat.get(key))
            months = solve_months(amount, base)
            # 基数单元格已在模板里写好（B15=2000 等），保持不动
            ws[f"{MONTH_COLS[i]}{r}"] = months
            ws[f"{TYPE_COLS[i]}{r}"] = type_val
            if use_num:
                ws[f"{NUM_COLS[i]}{r}"] = 1 if amount > 0 else 0

    # 继续教育（第 20 行）：F20=B20*12 —— 写月额度
    for i, y in enumerate(years):
        flat = by_year[y]["flat"]
        ws[f"{YEAR_COLS[i]}20"] = to_float(flat.get("继续教育")) / 12.0

    # 第 21 行 大病医疗：公式引用统一的 $E$45（自付金额），三个年度共用。
    # 采集数据里大病均为 0，保持 E45=0 即可；如某年有值，模板本身限制只能统一填。
    # 基本减除费用（第 10 行）B/G/L 模板已写死 60000，不动。

    # ---- 5. 第 40 行 已预缴税额 / 第 41 行 已退税额
    for i, y in enumerate(years):
        summary = by_year[y]["summary"]
        paid = to_float(summary.get("已缴税额"))
        ws[f"{PREPAID_COLS[i]}40"] = paid
        ws[f"{REFUND_COLS[i]}41"] = 0          # 无已退记录，留 0（第41行 R41 备注说明填负数）

    # ---- 6. 标题加上纳税人姓名，R1 记录采集时间
    抓取时间 = data.get("抓取时间", "")
    if name:
        ws["A1"] = f"个人所得税退税计算表 — {name}（{'-'.join(str(y) for y in years)}）"
    if 抓取时间:
        ws["R1"] = f"数据采集时间：{抓取时间}"

    # ---- 保存
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"✓ 已生成：{out_path}")

    # ---- 控制台校验：用同样的公式逻辑算一遍，打印关键结果
    print("\n=== 校验（按模板公式复算）===")
    for i, y in enumerate(years):
        flat = by_year[y]["flat"]
        summary = by_year[y]["summary"]
        收入 = to_float(flat.get("工资薪金"))
        扣除 = (
            60000
            + to_float(flat.get("基本养老保险"))
            + to_float(flat.get("基本医疗保险"))
            + to_float(flat.get("失业保险"))
            + to_float(flat.get("住房公积金"))
            + to_float(flat.get("子女教育"))
            + to_float(flat.get("赡养老人"))
            + to_float(flat.get("住房贷款利息"))
            + to_float(flat.get("住房租金"))
            + to_float(flat.get("继续教育"))
            + to_float(flat.get("商业健康险"))
            + to_float(flat.get("税延养老保险"))
            + to_float(flat.get("个人养老金"))
            + to_float(flat.get("准予扣除的捐赠额"))
            + to_float(flat.get("其他"))
            + to_float(flat.get("3岁以下婴幼儿照护"))
            + to_float(flat.get("年金"))
        )
        应税 = max(收入 - 扣除, 0)
        税率, 速扣 = rate_and_quick(应税)
        应纳税额 = round(应税 * 税率 - 速扣, 2)
        已缴 = to_float(summary.get("已缴税额"))
        退补 = round(应纳税额 - 已缴, 2)
        print(
            f"{y}年：收入 {收入:,.2f} | 扣除 {扣除:,.2f} | "
            f"应税 {应税:,.2f} | 应纳税额 {应纳税额:,.2f} | "
            f"已缴 {已缴:,.2f} | 退(补) {退补:,.2f}"
        )
        print(
            f"      申报原值：应纳税额 {to_float(summary.get('应纳税额')):,.2f}  "
            f"应纳税所得额 {to_float(summary.get('应纳税所得额')):,.2f}"
        )


def rate_and_quick(taxable):
    """与模板第 30/31 行一致的 7 级综合所得年度税率表。"""
    brackets = [
        (36000, 0.03, 0),
        (144000, 0.10, 2520),
        (300000, 0.20, 16920),
        (420000, 0.25, 31920),
        (660000, 0.30, 52920),
        (960000, 0.35, 85920),
        (float("inf"), 0.45, 181920),
    ]
    for cap, rate, quick in brackets:
        if taxable <= cap:
            return rate, quick
    return 0.45, 181920


if __name__ == "__main__":
    j = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_JSON
    t = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_TPL
    o = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_OUT
    main(j, t, o)
